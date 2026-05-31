#!/usr/bin/env python3
"""Deterministic helper for publication-ready LaTeX table workflows."""

from __future__ import annotations

import argparse
import contextlib
import csv
import hashlib
import io
import json
import os
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any
from datetime import date, datetime, time


HIGHER_IS_BETTER = {
    "accuracy",
    "acc",
    "f1",
    "auc",
    "auroc",
    "auprc",
    "precision",
    "recall",
    "score",
    "performance",
    "success",
    "bleu",
    "rouge",
    "meteor",
    "cider",
    "map",
    "ndcg",
}

LOWER_IS_BETTER = {
    "loss",
    "error",
    "err",
    "mae",
    "mse",
    "rmse",
    "mape",
    "perplexity",
    "ppl",
    "wer",
    "cer",
    "latency",
    "time",
    "memory",
    "cost",
    "params",
    "flops",
}

COMMON_PACKAGES = [
    "booktabs",
    "tabularx",
    "array",
    "adjustbox",
    "xcolor",
    "colortbl",
    "siunitx",
    "multirow",
    "makecell",
    "caption",
    "subcaption",
]

DEFAULT_REQUIRED_PACKAGES = ["booktabs", "tabularx", "array"]


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8-sig", errors="replace")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(read_text(path))


def stable_fingerprint(columns: list[str], rows: list[list[str]]) -> str:
    return hashlib.sha256(
        json.dumps({"columns": columns, "rows": rows}, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()


def parse_key_value_rules(rules: str | None) -> dict[str, str]:
    parsed: dict[str, str] = {}
    if not rules:
        return parsed
    for item in re.split(r"[,;]", rules):
        item = item.strip()
        if not item:
            continue
        if "=" not in item:
            raise ValueError(f"Metric direction rule must use name=direction: {item}")
        name, direction = [part.strip() for part in item.split("=", 1)]
        direction = direction.lower()
        if direction not in {"higher", "lower", "none"}:
            raise ValueError(f"Unsupported direction for {name}: {direction}")
        parsed[name.lower()] = direction
    return parsed


def load_generation_config(path: str | None) -> dict[str, Any]:
    if not path:
        return {}
    config = load_json(Path(path))
    if not isinstance(config, dict):
        raise ValueError("Generation config must be a JSON object.")
    return config


def config_metric_rules(config: dict[str, Any]) -> dict[str, str]:
    raw = config.get("metric_directions", {})
    if raw is None:
        return {}
    if not isinstance(raw, dict):
        raise ValueError("config.metric_directions must be an object.")
    parsed: dict[str, str] = {}
    for name, direction in raw.items():
        direction_text = str(direction).lower()
        if direction_text not in {"higher", "lower", "none"}:
            raise ValueError(f"Unsupported direction for {name}: {direction}")
        parsed[str(name).lower()] = direction_text
    return parsed


def column_indices(columns: list[str], names: list[str]) -> list[int]:
    lookup = {column.lower(): index for index, column in enumerate(columns)}
    indices: list[int] = []
    missing: list[str] = []
    for name in names:
        key = str(name).lower()
        if key not in lookup:
            missing.append(str(name))
        else:
            indices.append(lookup[key])
    if missing:
        raise ValueError(f"Unknown column(s): {', '.join(missing)}")
    return indices


def config_string_list(config: dict[str, Any], key: str) -> list[str]:
    raw = config.get(key, [])
    if raw is None:
        return []
    if not isinstance(raw, list):
        raise ValueError(f"config.{key} must be a list.")
    return [str(item) for item in raw]


def package_is_loaded(packages: dict[str, Any], package: str) -> bool:
    return package in packages


def latex_escape(value: Any) -> str:
    text = "" if value is None else str(value)
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
        "~": r"\textasciitilde{}",
        "^": r"\textasciicircum{}",
    }
    return "".join(replacements.get(char, char) for char in text)


def latex_unescape(value: str) -> str:
    text = value
    replacements = {
        r"\&": "&",
        r"\%": "%",
        r"\$": "$",
        r"\#": "#",
        r"\_": "_",
        r"\{": "{",
        r"\}": "}",
        r"\textbackslash{}": "\\",
        r"\textasciitilde{}": "~",
        r"\textasciicircum{}": "^",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def normalize_cell(value: Any) -> str:
    text = "" if value is None else str(value)
    text = latex_to_plain_text(text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def latex_to_plain_text(value: str) -> str:
    text = value.strip()
    previous = None
    while previous != text:
        previous = text
        text = re.sub(r"\\(?:textbf|underline|emph|mathbf|cellcolor)(?:\[[^\]]*\])?\{([^{}]*)\}", r"\1", text)
        text = re.sub(r"\\multicolumn\{[^{}]*\}\{[^{}]*\}\{([^{}]*)\}", r"\1", text)
        text = re.sub(r"\\multirow\{[^{}]*\}\{[^{}]*\}\{([^{}]*)\}", r"\1", text)
    text = re.sub(r"\\(?:toprule|midrule|bottomrule|hline|cmidrule)(?:\([^)]*\))?(?:\{[^{}]*\})?", "", text)
    text = re.sub(r"\\[a-zA-Z]+\*?(?:\[[^\]]*\])?", "", text)
    text = text.replace("{", "").replace("}", "")
    return latex_unescape(text)


def split_latex_rows(body: str) -> list[str]:
    rows: list[str] = []
    current: list[str] = []
    index = 0
    while index < len(body):
        if body[index : index + 2] == "\\\\":
            rows.append("".join(current))
            current = []
            index += 2
            continue
        current.append(body[index])
        index += 1
    tail = "".join(current).strip()
    if tail:
        rows.append(tail)
    return rows


def split_latex_cells(row: str) -> list[str]:
    cells: list[str] = []
    current: list[str] = []
    brace_depth = 0
    escaped = False
    for char in row:
        if escaped:
            current.append(char)
            escaped = False
            continue
        if char == "\\":
            current.append(char)
            escaped = True
            continue
        if char == "{":
            brace_depth += 1
        elif char == "}" and brace_depth:
            brace_depth -= 1
        if char == "&" and brace_depth == 0:
            cells.append("".join(current).strip())
            current = []
        else:
            current.append(char)
    cells.append("".join(current).strip())
    return cells


def strip_environment_arguments(body: str) -> str:
    text = body.lstrip()
    consumed = True
    while consumed:
        consumed = False
        text = text.lstrip()
        if text.startswith("["):
            depth = 0
            for index, char in enumerate(text):
                if char == "[":
                    depth += 1
                elif char == "]":
                    depth -= 1
                    if depth == 0:
                        text = text[index + 1 :]
                        consumed = True
                        break
        text = text.lstrip()
        if text.startswith("{"):
            depth = 0
            for index, char in enumerate(text):
                if char == "{":
                    depth += 1
                elif char == "}":
                    depth -= 1
                    if depth == 0:
                        text = text[index + 1 :]
                        consumed = True
                        break
    return text


def extract_tabular_cells(text: str) -> tuple[list[str], list[list[str]], list[str]]:
    warnings: list[str] = []
    begin_match = re.search(r"\\begin\{(tabularx|tabular\*|tabular)\}", text)
    if not begin_match:
        raise ValueError("No standard tabular, tabularx, or tabular* environment found.")
    environment = begin_match.group(1)
    end_match = re.search(rf"\\end\{{{re.escape(environment)}\}}", text[begin_match.end() :], flags=re.DOTALL)
    if not end_match:
        raise ValueError(f"No matching end environment found for {environment}.")
    body = text[begin_match.end() : begin_match.end() + end_match.start()]
    body = strip_environment_arguments(body)
    if "\\multirow" in body or "\\multicolumn" in body:
        warnings.append("Detected multirow/multicolumn; parser preserves visible cell text but cannot infer full spanning semantics.")
    cleaned_rows: list[list[str]] = []
    for raw_row in split_latex_rows(body):
        row = re.sub(r"\\(?:toprule|midrule|bottomrule|hline)\b", "", raw_row)
        row = re.sub(r"\\cmidrule(?:\([^)]*\))?\{[^{}]*\}", "", row)
        row = row.strip()
        if not row:
            continue
        cells = [normalize_cell(cell) for cell in split_latex_cells(row)]
        if any(cell for cell in cells):
            cleaned_rows.append(cells)
    if not cleaned_rows:
        raise ValueError("The tabular environment did not contain parseable rows.")
    width = max(len(row) for row in cleaned_rows)
    padded = [row + [""] * (width - len(row)) for row in cleaned_rows]
    return padded[0], padded[1:], warnings


def read_template_family(tex_path: Path) -> tuple[str, list[dict[str, Any]]]:
    seen: set[Path] = set()
    files: list[dict[str, Any]] = []

    def visit(path: Path, depth: int) -> str:
        resolved = path.resolve()
        if resolved in seen or depth > 1 or not resolved.exists():
            return ""
        seen.add(resolved)
        text = read_text(resolved)
        files.append({"path": str(resolved), "sha256": sha256_file(resolved), "bytes": resolved.stat().st_size})
        chunks = [text]
        for match in re.finditer(r"\\(?:input|include)\{([^{}]+)\}", text):
            raw = match.group(1).strip()
            child = (resolved.parent / raw)
            if child.suffix == "":
                child = child.with_suffix(".tex")
            chunks.append(visit(child, depth + 1))
        for match in re.finditer(r"\\documentclass(?:\[[^\]]*\])?\{([^{}]+)\}", text):
            cls = resolved.parent / f"{match.group(1).strip()}.cls"
            chunks.append(visit(cls, depth + 1))
        for match in re.finditer(r"\\usepackage(?:\[[^\]]*\])?\{([^{}]+)\}", text):
            for package in [part.strip() for part in match.group(1).split(",")]:
                style = resolved.parent / f"{package}.sty"
                chunks.append(visit(style, depth + 1))
        return "\n".join(chunks)

    combined = visit(tex_path, 0)
    return combined, files


def analyze_template(args: argparse.Namespace) -> None:
    tex_path = Path(args.tex)
    text, template_files = read_template_family(tex_path)
    class_match = re.search(r"\\documentclass(?:\[(?P<options>[^\]]*)\])?\{(?P<class>[^{}]+)\}", text)
    options = []
    doc_class = "unknown"
    if class_match:
        doc_class = class_match.group("class")
        options = [part.strip() for part in (class_match.group("options") or "").split(",") if part.strip()]
    packages: dict[str, dict[str, Any]] = {}
    for pkg_match in re.finditer(r"\\usepackage(?:\[(?P<options>[^\]]*)\])?\{(?P<packages>[^{}]+)\}", text):
        pkg_options = [part.strip() for part in (pkg_match.group("options") or "").split(",") if part.strip()]
        for package in [part.strip() for part in pkg_match.group("packages").split(",")]:
            packages[package] = {"options": pkg_options}
    lower_class = doc_class.lower()
    option_lowers = [option.lower() for option in options]
    is_two_column = (
        "twocolumn" in [option.lower() for option in options]
        or "sigconf" in [option.lower() for option in options]
        or lower_class in {"acmart", "ieeetran"}
        or "\\twocolumn" in text
    )
    if "onecolumn" in option_lowers or "\\onecolumn" in text:
        is_two_column = False
    profile = {
        "source": {"path": str(tex_path), "sha256": sha256_file(tex_path)},
        "template_files": template_files,
        "document_class": doc_class,
        "class_options": options,
        "is_two_column": is_two_column,
        "packages": packages,
        "known_table_packages": {name: packages.get(name) for name in COMMON_PACKAGES if name in packages},
        "color_available": "xcolor" in packages or "colortbl" in packages,
        "xcolor_table_available": "colortbl" in packages
        or ("xcolor" in packages and "table" in packages.get("xcolor", {}).get("options", [])),
        "recommended_width_macro": r"\textwidth" if is_two_column else r"\columnwidth",
        "existing_table_star_usage": bool(re.search(r"\\begin\{table\*\}", text)),
        "existing_resizebox_usage": bool(re.search(r"\\resizebox\b", text)),
        "caption_package_loaded": "caption" in packages or "subcaption" in packages,
        "notes": [],
    }
    if lower_class in {"acmart", "ieeetran"} and not is_two_column:
        profile["notes"].append("Class is often two-column; verify final mode if template options are incomplete.")
    write_json(Path(args.out), profile)


def rows_to_table(rows: list[list[Any]], header_row: int = 1) -> tuple[list[str], list[list[str]]]:
    if not rows:
        raise ValueError("Input table is empty.")
    if header_row < 1 or header_row > len(rows):
        raise ValueError(f"header_row must be between 1 and {len(rows)}.")
    width = max(len(row) for row in rows)
    normalized = [[str(cell) for cell in row] + [""] * (width - len(row)) for row in rows]
    header_index = header_row - 1
    columns = normalized[header_index]
    body = normalized[header_index + 1 :]
    return columns, body


def read_delimited_table(path: Path, delimiter: str, header_row: int = 1) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=delimiter))
    return rows_to_table(rows, header_row)


def read_json_table(path: Path, header_row: int = 1) -> tuple[list[str], list[list[str]]]:
    payload = json.loads(read_text(path))
    if isinstance(payload, dict) and "columns" in payload and "rows" in payload:
        return [str(cell) for cell in payload["columns"]], [[str(cell) for cell in row] for row in payload["rows"]]
    if isinstance(payload, list) and payload and isinstance(payload[0], dict):
        columns: list[str] = []
        for item in payload:
            for key in item.keys():
                if key not in columns:
                    columns.append(str(key))
        rows = [[str(item.get(column, "")) for column in columns] for item in payload]
        return columns, rows
    if isinstance(payload, list) and all(isinstance(row, list) for row in payload):
        return rows_to_table(payload, header_row)
    raise ValueError("JSON input must be {columns, rows}, a list of objects, or a list of rows.")


def read_xlsx_table(path: Path, sheet_name: str | None = None, header_row: int = 1) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for .xlsx input.") from exc
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}. Available sheets: {', '.join(workbook.sheetnames)}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    def format_xlsx_cell(cell: Any) -> str:
        value = cell.value
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ")
        if isinstance(value, (date, time)):
            return value.isoformat()
        if isinstance(value, (int, float)):
            number_format = str(getattr(cell, "number_format", "") or "")
            if "%" in number_format:
                decimals = 0
                match = re.search(r"\.([0#]+)", number_format)
                if match:
                    decimals = len(match.group(1))
                return f"{float(value) * 100:.{decimals}f}%"
            match = re.search(r"\.([0#]+)", number_format)
            if match:
                decimals = len(match.group(1))
                return f"{float(value):.{decimals}f}"
            return str(value)
        return str(value)

    values = [[format_xlsx_cell(cell) for cell in row] for row in sheet.iter_rows()]
    values = [row for row in values if any(cell != "" for cell in row)]
    return rows_to_table(values, header_row)


def ingest(args: argparse.Namespace) -> None:
    source_path = Path(args.input)
    suffix = source_path.suffix.lower()
    warnings: list[str] = []
    if suffix == ".csv":
        columns, rows = read_delimited_table(source_path, ",", args.header_row)
        source_type = "csv"
    elif suffix == ".tsv":
        columns, rows = read_delimited_table(source_path, "\t", args.header_row)
        source_type = "tsv"
    elif suffix == ".xlsx":
        columns, rows = read_xlsx_table(source_path, args.sheet, args.header_row)
        source_type = "xlsx"
    elif suffix == ".json":
        columns, rows = read_json_table(source_path, args.header_row)
        source_type = "json"
    elif suffix == ".tex":
        columns, rows, warnings = extract_tabular_cells(read_text(source_path))
        source_type = "latex"
    else:
        raise ValueError(f"Unsupported input type: {source_path.suffix}")
    data = {
        "version": 1,
        "source": {"path": str(source_path), "type": source_type, "sha256": sha256_file(source_path)},
        "ingest_options": {"sheet": args.sheet, "header_row": args.header_row},
        "columns": [str(cell) for cell in columns],
        "rows": [[str(cell) for cell in row] for row in rows],
        "column_count": len(columns),
        "row_count": len(rows),
        "raw_cell_fingerprint": stable_fingerprint([str(cell) for cell in columns], [[str(cell) for cell in row] for row in rows]),
        "warnings": warnings,
    }
    write_json(Path(args.out), data)


def metric_direction(name: str, overrides: dict[str, str] | None = None) -> str | None:
    overrides = overrides or {}
    if name.lower() in overrides:
        direction = overrides[name.lower()]
        return None if direction == "none" else direction
    key = re.sub(r"[^a-z0-9]+", " ", name.lower()).strip()
    tokens = set(key.split())
    for item in HIGHER_IS_BETTER:
        if item in tokens or item in key:
            return "higher"
    for item in LOWER_IS_BETTER:
        if item in tokens or item in key:
            return "lower"
    return None


def parse_number(value: str) -> float | None:
    text = normalize_cell(value).replace(",", "")
    match = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def compute_highlights(
    columns: list[str],
    rows: list[list[str]],
    overrides: dict[str, str] | None = None,
    second_best: bool = True,
    compare_by_indices: list[int] | None = None,
) -> tuple[dict[tuple[int, int], str], list[dict[str, Any]]]:
    highlights: dict[tuple[int, int], str] = {}
    decisions: list[dict[str, Any]] = []
    compare_by_indices = compare_by_indices or []
    groups: dict[tuple[str, ...], list[int]] = {}
    if compare_by_indices:
        for row_index, row in enumerate(rows):
            key = tuple(row[index] if index < len(row) else "" for index in compare_by_indices)
            groups.setdefault(key, []).append(row_index)
    else:
        groups[("__all__",)] = list(range(len(rows)))
    for col_index, column in enumerate(columns):
        if col_index in compare_by_indices:
            decisions.append({"column": column, "direction": None, "reason": "compare_by_column"})
            continue
        override_direction = (overrides or {}).get(column.lower())
        direction = metric_direction(column, overrides)
        if direction is None:
            reason = "disabled_by_override" if override_direction == "none" else "ambiguous"
            decisions.append({"column": column, "direction": None, "reason": reason})
            continue
        group_decisions: list[dict[str, Any]] = []
        for group_key, row_indices in groups.items():
            values: list[tuple[int, float]] = []
            for row_index in row_indices:
                row = rows[row_index]
                if col_index >= len(row):
                    continue
                number = parse_number(row[col_index])
                if number is not None:
                    values.append((row_index, number))
            if len(values) < 2:
                group_decisions.append({"group": list(group_key), "reason": "not_enough_numeric_values"})
                continue
            reverse = direction == "higher"
            unique_sorted = sorted({value for _, value in values}, reverse=reverse)
            if not unique_sorted:
                group_decisions.append({"group": list(group_key), "reason": "no_numeric_values"})
                continue
            best = unique_sorted[0]
            second = unique_sorted[1] if second_best and len(unique_sorted) > 1 else None
            for row_index, value in values:
                if value == best:
                    highlights[(row_index, col_index)] = "best"
                elif second is not None and value == second:
                    highlights[(row_index, col_index)] = "second"
            group_decisions.append(
                {
                    "group": list(group_key),
                    "best": best,
                    "second": second,
                    "numeric_values": len(values),
                }
            )
        decisions.append({"column": column, "direction": direction, "groups": group_decisions})
    return highlights, decisions


def format_cell(value: str, highlight: str | None = None) -> str:
    escaped = latex_escape(value)
    if highlight == "best":
        return rf"\textbf{{{escaped}}}"
    if highlight == "second":
        return rf"\underline{{{escaped}}}"
    return escaped


def column_spec(column_count: int, use_tabularx: bool) -> str:
    if column_count <= 0:
        return "l"
    if use_tabularx:
        return "l" + " ".join([r">{\centering\arraybackslash}X"] * (column_count - 1))
    return "l" + ("".join(["c"] * (column_count - 1)))


def build_column_group_header(columns: list[str], config: dict[str, Any]) -> tuple[str | None, list[str]]:
    raw_groups = config.get("column_groups", [])
    if not raw_groups:
        return None, []
    if not isinstance(raw_groups, list):
        raise ValueError("config.column_groups must be a list.")
    spans: dict[int, tuple[int, str]] = {}
    cmidrules: list[str] = []
    covered: set[int] = set()
    for group in raw_groups:
        if not isinstance(group, dict):
            raise ValueError("Each column group must be an object.")
        label = str(group.get("label", "")).strip()
        names = group.get("columns", [])
        if not label or not isinstance(names, list) or not names:
            raise ValueError("Each column group needs a non-empty label and columns list.")
        indices = column_indices(columns, [str(name) for name in names])
        expected = list(range(min(indices), max(indices) + 1))
        if indices != expected:
            raise ValueError(f"Column group must be contiguous: {label}")
        if covered.intersection(indices):
            raise ValueError(f"Column group overlaps another group: {label}")
        covered.update(indices)
        spans[indices[0]] = (len(indices), label)
        cmidrules.append(rf"\cmidrule(lr){{{indices[0] + 1}-{indices[-1] + 1}}}")
    cells: list[str] = []
    index = 0
    while index < len(columns):
        if index in spans:
            span, label = spans[index]
            cells.append(rf"\multicolumn{{{span}}}{{c}}{{{latex_escape(label)}}}")
            index += span
        else:
            cells.append("")
            index += 1
    return " & ".join(cells) + r" \\", cmidrules


def build_table_notes(config: dict[str, Any]) -> list[str]:
    notes = config_string_list(config, "notes")
    if not notes:
        return []
    joined = " ".join(latex_escape(note) for note in notes)
    return [r"\vspace{2pt}", rf"{{\footnotesize {joined}}}"]


def table_quality_findings(table_text: str, template: dict[str, Any] | None = None) -> list[dict[str, str]]:
    template = template or {}
    findings: list[dict[str, str]] = []
    if r"\resizebox" in table_text:
        findings.append({"severity": "warn", "code": "resizebox-used", "message": "resizebox can hide unreadable tables; prefer structural fitting."})
    if r"\tiny" in table_text or r"\scriptsize" in table_text:
        findings.append({"severity": "warn", "code": "very-small-font", "message": "Very small table fonts may be unreadable in print."})
    if r"\cellcolor" in table_text and not template.get("xcolor_table_available"):
        findings.append({"severity": "error", "code": "cellcolor-without-xcolor-table", "message": "cellcolor requires colortbl or xcolor with table option."})
    if r"\begin{table*}" in table_text and not template.get("is_two_column"):
        findings.append({"severity": "warn", "code": "table-star-in-one-column", "message": "table* is usually unnecessary in one-column templates."})
    if r"\begin{tabularx}" in table_text:
        packages = template.get("packages", {})
        if not package_is_loaded(packages, "tabularx"):
            findings.append({"severity": "info", "code": "requires-tabularx", "message": "Generated table requires \\usepackage{tabularx}."})
        if not package_is_loaded(packages, "array"):
            findings.append({"severity": "info", "code": "requires-array", "message": "Generated table requires \\usepackage{array}."})
    if r"\toprule" in table_text and not package_is_loaded(template.get("packages", {}), "booktabs"):
        findings.append({"severity": "info", "code": "requires-booktabs", "message": "Generated table requires \\usepackage{booktabs}."})
    return findings


def generate(args: argparse.Namespace) -> None:
    data = load_json(Path(args.data))
    template = load_json(Path(args.template))
    config = load_generation_config(args.config)
    columns = [str(column) for column in data["columns"]]
    rows = [[str(cell) for cell in row] for row in data["rows"]]
    column_count = len(columns)
    is_two_column = bool(template.get("is_two_column"))
    if args.wide == "yes":
        use_wide_float = True
    elif args.wide == "no":
        use_wide_float = False
    else:
        use_wide_float = is_two_column and column_count >= 7
    use_tabularx = column_count >= 5 or any(len(cell) > 18 for cell in columns + [cell for row in rows for cell in row])
    table_env = "table*" if use_wide_float else "table"
    width_macro = r"\textwidth" if use_wide_float else r"\linewidth"
    metric_overrides = config_metric_rules(config)
    metric_overrides.update(parse_key_value_rules(args.metric_directions))
    compare_by_names = config_string_list(config, "compare_by")
    compare_by = column_indices(columns, compare_by_names) if compare_by_names else []
    row_group_names = config_string_list(config, "row_group_by")
    row_group_indices = column_indices(columns, row_group_names) if row_group_names else []
    if args.no_highlight:
        highlights: dict[tuple[int, int], str] = {}
        highlight_decisions: list[dict[str, Any]] = []
    else:
        highlights, highlight_decisions = compute_highlights(
            columns,
            rows,
            metric_overrides,
            not args.no_second_best,
            compare_by,
        )
    required_packages = ["booktabs"]
    if use_tabularx:
        required_packages.extend(["tabularx", "array"])
    required_packages = [pkg for pkg in required_packages if pkg not in template.get("packages", {})]
    lines: list[str] = []
    if required_packages:
        lines.append("% Required packages: " + ", ".join(required_packages))
    lines.append(f"% Source fingerprint: {data.get('raw_cell_fingerprint', 'unknown')}")
    lines.append(rf"\begin{{{table_env}}}[t]")
    lines.append(r"\centering")
    if column_count >= 8 or args.style == "compact":
        lines.append(r"\small")
        lines.append(r"\setlength{\tabcolsep}{4pt}")
    lines.append(rf"\caption{{{latex_escape(args.caption)}}}")
    lines.append(rf"\label{{{latex_escape(args.label)}}}")
    spec = column_spec(column_count, use_tabularx)
    if use_tabularx:
        lines.append(rf"\begin{{tabularx}}{{{width_macro}}}{{{spec}}}")
    else:
        lines.append(rf"\begin{{tabular}}{{{spec}}}")
    lines.append(r"\toprule")
    group_header, cmidrules = build_column_group_header(columns, config)
    if group_header:
        lines.append(group_header)
        if cmidrules:
            lines.append(" ".join(cmidrules))
    lines.append(" & ".join(latex_escape(column) for column in columns) + r" \\")
    lines.append(r"\midrule")
    previous_row_group: tuple[str, ...] | None = None
    for row_index, row in enumerate(rows):
        if row_group_indices:
            current_row_group = tuple(row[index] if index < len(row) else "" for index in row_group_indices)
            if previous_row_group is not None and current_row_group != previous_row_group:
                lines.append(r"\addlinespace")
            previous_row_group = current_row_group
        padded = row + [""] * (column_count - len(row))
        cells = [
            format_cell(cell, highlights.get((row_index, col_index)))
            for col_index, cell in enumerate(padded[:column_count])
        ]
        lines.append(" & ".join(cells) + r" \\")
    lines.append(r"\bottomrule")
    lines.append(r"\end{tabularx}" if use_tabularx else r"\end{tabular}")
    lines.extend(build_table_notes(config))
    lines.append(rf"\end{{{table_env}}}")
    table_text = "\n".join(lines) + "\n"
    Path(args.out).write_text(table_text, encoding="utf-8")
    manifest = {
        "ok": True,
        "table": str(Path(args.out)),
        "source_fingerprint": data.get("raw_cell_fingerprint"),
        "source": data.get("source"),
        "template_source": template.get("source"),
        "layout": {
            "table_environment": table_env,
            "tabular_environment": "tabularx" if use_tabularx else "tabular",
            "width_macro": width_macro,
            "column_spec": spec,
            "style": args.style,
            "column_groups": config.get("column_groups", []),
            "compare_by": compare_by_names,
            "row_group_by": row_group_names,
        },
        "config": str(Path(args.config)) if args.config else None,
        "required_packages": required_packages,
        "highlight_decisions": highlight_decisions,
        "quality_findings": table_quality_findings(table_text, template),
    }
    manifest_path = Path(args.manifest) if args.manifest else Path(args.out).with_suffix(Path(args.out).suffix + ".manifest.json")
    write_json(manifest_path, manifest)


def validate(args: argparse.Namespace) -> None:
    data = load_json(Path(args.data))
    columns = [normalize_cell(column) for column in data["columns"]]
    expected_rows = [[normalize_cell(cell) for cell in row] for row in data["rows"]]
    table_text = read_text(Path(args.table))
    actual_columns, actual_rows, warnings = extract_tabular_cells(table_text)
    actual_columns = [normalize_cell(cell) for cell in actual_columns]
    actual_rows = [[normalize_cell(cell) for cell in row] for row in actual_rows]
    if actual_columns[: len(columns)] != columns and actual_rows and actual_rows[0][: len(columns)] == columns:
        warnings = [warning for warning in warnings if not warning.startswith("Detected multirow/multicolumn")]
        actual_columns = actual_rows[0]
        actual_rows = actual_rows[1:]
    actual_rows = [row for row in actual_rows if not (len(row) == 1 and row[0] and row[0] not in {expected[0] for expected in expected_rows if expected})]
    mismatches: list[dict[str, Any]] = []
    if actual_columns[: len(columns)] != columns:
        mismatches.append({"kind": "header", "expected": columns, "actual": actual_columns})
    if len(actual_rows) != len(expected_rows):
        mismatches.append({"kind": "row_count", "expected": len(expected_rows), "actual": len(actual_rows)})
    for row_index, expected_row in enumerate(expected_rows):
        if row_index >= len(actual_rows):
            continue
        actual_row = actual_rows[row_index]
        width = max(len(expected_row), len(actual_row))
        for col_index in range(width):
            expected = expected_row[col_index] if col_index < len(expected_row) else ""
            actual = actual_row[col_index] if col_index < len(actual_row) else ""
            if expected != actual:
                mismatches.append(
                    {
                        "kind": "cell",
                        "row": row_index,
                        "column": col_index,
                        "expected": expected,
                        "actual": actual,
                    }
                )
    report = {
        "ok": not mismatches,
        "mismatch_count": len(mismatches),
        "mismatches": mismatches,
        "warnings": warnings,
    }
    if args.report:
        write_json(Path(args.report), report)
    print(json.dumps(report, indent=2, ensure_ascii=False))
    if mismatches:
        raise SystemExit(1)


def audit(args: argparse.Namespace) -> None:
    template = load_json(Path(args.template)) if args.template else {}
    table_text = read_text(Path(args.table))
    findings = table_quality_findings(table_text, template)
    error_count = sum(1 for finding in findings if finding["severity"] == "error")
    report = {"ok": error_count == 0, "finding_count": len(findings), "findings": findings}
    if args.report:
        write_json(Path(args.report), report)
    print(json.dumps(report, indent=2, ensure_ascii=False))
    if error_count:
        raise SystemExit(1)


def analyze_rendered_image(path: Path) -> dict[str, Any]:
    try:
        from PIL import Image
    except ImportError:
        return {"path": str(path), "ok": None, "warning": "Pillow is not available; image analysis skipped."}
    image = Image.open(path).convert("L")
    width, height = image.size
    pixels = image.load()
    min_x, min_y = width, height
    max_x, max_y = -1, -1
    for y in range(height):
        for x in range(width):
            if pixels[x, y] < 245:
                min_x = min(min_x, x)
                min_y = min(min_y, y)
                max_x = max(max_x, x)
                max_y = max(max_y, y)
    findings: list[dict[str, str]] = []
    if max_x == -1:
        findings.append({"severity": "error", "code": "blank-preview", "message": "Rendered preview appears blank."})
        content_box = None
        content_ratio = 0.0
    else:
        box_width = max_x - min_x + 1
        box_height = max_y - min_y + 1
        content_box = {"x": min_x, "y": min_y, "width": box_width, "height": box_height}
        content_ratio = (box_width * box_height) / float(width * height)
        if box_width / width > 0.96:
            findings.append({"severity": "warn", "code": "near-page-width", "message": "Rendered content nearly spans the full page width; inspect for clipping."})
        if content_ratio < 0.001:
            findings.append({"severity": "warn", "code": "tiny-rendered-content", "message": "Rendered content is extremely small on the page."})
    return {
        "path": str(path),
        "ok": not any(finding["severity"] == "error" for finding in findings),
        "image_size": {"width": width, "height": height},
        "content_box": content_box,
        "content_ratio": content_ratio,
        "findings": findings,
    }


def find_tectonic() -> str | None:
    candidates = []
    env_path = os.environ.get("TECTONIC")
    if env_path:
        candidates.append(Path(env_path))
    candidates.extend(
        [
            Path(r"C:\Users\target\.codex\plugins\cache\target-local-plugins\latex-tectonic\0.1.0\bin\tectonic.exe"),
            Path(r"C:\Users\target\.codex\plugins\cache\target-local-plugins\latex-tectonic\0.1.0\bin\tectonic"),
        ]
    )
    for candidate in candidates:
        if candidate.exists():
            return str(candidate)
    return shutil.which("tectonic")


def build_preview_document(paper_tex: Path, table_tex: Path, outdir: Path) -> Path:
    paper = read_text(paper_tex)
    class_match = re.search(r"\\documentclass(?:\[[^\]]*\])?\{[^{}]+\}", paper)
    document_class = class_match.group(0) if class_match else r"\documentclass{article}"
    package_lines = re.findall(r"\\usepackage(?:\[[^\]]*\])?\{[^{}]+\}", paper)
    package_set = dict.fromkeys(package_lines)
    required = [r"\usepackage{booktabs}", r"\usepackage{tabularx}", r"\usepackage{array}"]
    for line in required:
        package_set.setdefault(line, None)
    preview_tex = outdir / "preview_table.tex"
    content = "\n".join(
        [
            document_class,
            *package_set.keys(),
            r"\begin{document}",
            read_text(table_tex),
            r"\end{document}",
            "",
        ]
    )
    preview_tex.write_text(content, encoding="utf-8")
    return preview_tex


def preview(args: argparse.Namespace) -> None:
    paper_tex = Path(args.paper).resolve()
    table_tex = Path(args.table).resolve()
    outdir = Path(args.outdir).resolve()
    outdir.mkdir(parents=True, exist_ok=True)
    preview_tex = build_preview_document(paper_tex, table_tex, outdir)
    tectonic = find_tectonic()
    if not tectonic:
        raise RuntimeError("Could not find Tectonic. Install it or set the TECTONIC environment variable.")
    command = [tectonic, "--outdir", str(outdir), str(preview_tex)]
    result = subprocess.run(command, cwd=str(outdir), text=True, capture_output=True)
    (outdir / "tectonic.stdout.txt").write_text(result.stdout, encoding="utf-8")
    (outdir / "tectonic.stderr.txt").write_text(result.stderr, encoding="utf-8")
    if result.returncode != 0:
        raise RuntimeError(f"Tectonic failed. See {outdir / 'tectonic.stderr.txt'}")
    pdf_path = outdir / "preview_table.pdf"
    render_outputs: list[str] = []
    pdftoppm = shutil.which("pdftoppm")
    if pdftoppm and pdf_path.exists():
        render_prefix = outdir / "preview_page"
        render = subprocess.run(
            [pdftoppm, "-png", "-r", "180", str(pdf_path), str(render_prefix)],
            cwd=str(outdir),
            text=True,
            capture_output=True,
        )
        (outdir / "pdftoppm.stdout.txt").write_text(render.stdout, encoding="utf-8")
        (outdir / "pdftoppm.stderr.txt").write_text(render.stderr, encoding="utf-8")
        if render.returncode == 0:
            render_outputs = [str(path) for path in sorted(outdir.glob("preview_page-*.png"))]
    image_analysis = [analyze_rendered_image(Path(path)) for path in render_outputs]
    summary = {
        "ok": True,
        "tex": str(preview_tex),
        "pdf": str(pdf_path),
        "rendered_pages": render_outputs,
        "image_analysis": image_analysis,
        "log_findings": [
            line.strip()
            for line in (result.stdout + "\n" + result.stderr).splitlines()
            if "Overfull \\hbox" in line or "Underfull \\hbox" in line or "LaTeX Warning" in line
        ],
        "note": "Inspect rendered pages manually for width, font size, and visual clutter.",
    }
    if args.report:
        write_json(Path(args.report), summary)
    print(json.dumps(summary, indent=2, ensure_ascii=False))


def build(args: argparse.Namespace) -> None:
    outdir = Path(args.outdir).resolve()
    outdir.mkdir(parents=True, exist_ok=True)
    template_path = outdir / "template_profile.json"
    data_path = outdir / "table_data.json"
    table_path = outdir / "table.tex"
    manifest_path = outdir / "generation_manifest.json"
    validation_path = outdir / "validation.json"
    audit_path = outdir / "audit.json"
    preview_path = outdir / "preview"
    preview_report_path = outdir / "preview.json"
    summary_path = outdir / "build_summary.json"

    analyze_template(argparse.Namespace(tex=args.paper, out=str(template_path)))
    ingest(
        argparse.Namespace(
            input=args.input,
            out=str(data_path),
            sheet=args.sheet,
            header_row=args.header_row,
        )
    )
    generate(
        argparse.Namespace(
            data=str(data_path),
            template=str(template_path),
            caption=args.caption,
            label=args.label,
            out=str(table_path),
            config=args.config,
            manifest=str(manifest_path),
            metric_directions=args.metric_directions,
            no_highlight=args.no_highlight,
            no_second_best=args.no_second_best,
            wide=args.wide,
            style=args.style,
        )
    )
    with contextlib.redirect_stdout(io.StringIO()):
        validate(argparse.Namespace(data=str(data_path), table=str(table_path), report=str(validation_path)))
    with contextlib.redirect_stdout(io.StringIO()):
        audit(argparse.Namespace(table=str(table_path), template=str(template_path), report=str(audit_path)))
    preview_summary: dict[str, Any] | None = None
    if not args.skip_preview:
        with contextlib.redirect_stdout(io.StringIO()):
            preview(
                argparse.Namespace(
                    paper=args.paper,
                    table=str(table_path),
                    outdir=str(preview_path),
                    report=str(preview_report_path),
                )
            )
        preview_summary = load_json(preview_report_path)
    summary = {
        "ok": True,
        "artifacts": {
            "template_profile": str(template_path),
            "table_data": str(data_path),
            "table": str(table_path),
            "manifest": str(manifest_path),
            "validation": str(validation_path),
            "audit": str(audit_path),
            "preview": str(preview_report_path) if preview_summary else None,
        },
        "validation": load_json(validation_path),
        "audit": load_json(audit_path),
        "preview": preview_summary,
    }
    write_json(summary_path, summary)
    print(json.dumps(summary, indent=2, ensure_ascii=False))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    build_parser_cmd = subparsers.add_parser("build")
    build_parser_cmd.add_argument("--paper", required=True, help="Target manuscript/template .tex file.")
    build_parser_cmd.add_argument("--input", required=True, help="CSV/TSV/XLSX/JSON/LaTeX source data.")
    build_parser_cmd.add_argument("--caption", required=True)
    build_parser_cmd.add_argument("--label", required=True)
    build_parser_cmd.add_argument("--outdir", required=True, help="Directory for all generated artifacts.")
    build_parser_cmd.add_argument("--config", help="Optional JSON table spec.")
    build_parser_cmd.add_argument("--sheet", help="XLSX sheet name. Defaults to the active sheet.")
    build_parser_cmd.add_argument("--header-row", type=int, default=1)
    build_parser_cmd.add_argument("--metric-directions", help="Comma-separated rules such as Accuracy=higher,Loss=lower.")
    build_parser_cmd.add_argument("--no-highlight", action="store_true")
    build_parser_cmd.add_argument("--no-second-best", action="store_true")
    build_parser_cmd.add_argument("--wide", choices=["auto", "yes", "no"], default="auto")
    build_parser_cmd.add_argument("--style", choices=["journal", "blackwhite", "compact"], default="journal")
    build_parser_cmd.add_argument("--skip-preview", action="store_true", help="Skip Tectonic/PDF preview.")
    build_parser_cmd.set_defaults(func=build)

    analyze_parser = subparsers.add_parser("analyze-template")
    analyze_parser.add_argument("--tex", required=True)
    analyze_parser.add_argument("--out", required=True)
    analyze_parser.set_defaults(func=analyze_template)

    ingest_parser = subparsers.add_parser("ingest")
    ingest_parser.add_argument("--input", required=True)
    ingest_parser.add_argument("--out", required=True)
    ingest_parser.add_argument("--sheet", help="XLSX sheet name. Defaults to the active sheet.")
    ingest_parser.add_argument("--header-row", type=int, default=1, help="1-based header row for CSV/TSV/XLSX/row-list JSON.")
    ingest_parser.set_defaults(func=ingest)

    generate_parser = subparsers.add_parser("generate")
    generate_parser.add_argument("--data", required=True)
    generate_parser.add_argument("--template", required=True)
    generate_parser.add_argument("--caption", required=True)
    generate_parser.add_argument("--label", required=True)
    generate_parser.add_argument("--out", required=True)
    generate_parser.add_argument("--config", help="Optional JSON table spec for grouping, notes, and metric rules.")
    generate_parser.add_argument("--manifest", help="Optional path for generation manifest JSON.")
    generate_parser.add_argument("--metric-directions", help="Comma-separated rules such as Accuracy=higher,Loss=lower,Params=none.")
    generate_parser.add_argument("--no-highlight", action="store_true", help="Disable automatic best/second-best highlighting.")
    generate_parser.add_argument("--no-second-best", action="store_true", help="Only bold best values; do not underline second-best values.")
    generate_parser.add_argument("--wide", choices=["auto", "yes", "no"], default="auto", help="Control table* use.")
    generate_parser.add_argument("--style", choices=["journal", "blackwhite", "compact"], default="journal")
    generate_parser.set_defaults(func=generate)

    validate_parser = subparsers.add_parser("validate")
    validate_parser.add_argument("--data", required=True)
    validate_parser.add_argument("--table", required=True)
    validate_parser.add_argument("--report")
    validate_parser.set_defaults(func=validate)

    audit_parser = subparsers.add_parser("audit")
    audit_parser.add_argument("--table", required=True)
    audit_parser.add_argument("--template")
    audit_parser.add_argument("--report")
    audit_parser.set_defaults(func=audit)

    preview_parser = subparsers.add_parser("preview")
    preview_parser.add_argument("--paper", required=True)
    preview_parser.add_argument("--table", required=True)
    preview_parser.add_argument("--outdir", required=True)
    preview_parser.add_argument("--report")
    preview_parser.set_defaults(func=preview)

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    try:
        args.func(args)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
